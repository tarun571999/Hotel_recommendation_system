from flask import Flask,session, render_template, request, flash, redirect, url_for
from flask_debug import Debug
from werkzeug.utils import secure_filename


import pandas as pd

from openpyxl import load_workbook



from collections import Counter

import math
# import matplotlib.pyplot as plt
import seaborn as sns
import scipy.sparse
from scipy.sparse import csr_matrix
from scipy.sparse.linalg import svds
import numpy as np

#OTP PACKAGES
import random
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


app = Flask(__name__)
app.secret_key = "secret key"



def recommend_items(userID, pivot_df, preds_df, num_recommendations):
    # index starts at 0  
    user_idx = userID-1 
    # Get and sort the user's ratings
    sorted_user_ratings = pivot_df.iloc[user_idx].sort_values(ascending=False)
    #sorted_user_ratings
    sorted_user_predictions = preds_df.iloc[user_idx].sort_values(ascending=False)
    #sorted_user_predictions
    temp = pd.concat([sorted_user_ratings, sorted_user_predictions], axis=1,sort=False)
       
    temp.index.name = 'Recommended Items'
    temp.columns = ['user_ratings', 'user_predictions']
    #temp = temp.loc[temp.user_ratings == 0]   
    temp = temp.sort_values('user_predictions', ascending=False)
    print(temp.head())
    return temp

def recom(userID):
    
    user_df=pd.read_excel('input/new_userhotelrating_small.xlsx')
    
    print(user_df.head())   
    user_df = user_df.dropna()
    counts = user_df.user.value_counts()
    user_df_final = user_df[user_df.user.isin(counts[counts>=2].index)]
    train_data = user_df_final
    user_df_CF = train_data
    pivot_df = user_df_CF.pivot_table(index = 'user', columns ='product_code', values = 'rating').fillna(0)
    pivot_df['user_index'] = np.arange(0, pivot_df.shape[0], 1)
    pivot_df.set_index(['user_index'], inplace=True)
    U, sigma, Vt = svds(pivot_df, k = 10)
    sigma = np.diag(sigma)
    #Predicted ratings
    all_user_predicted_ratings = np.dot(np.dot(U, sigma), Vt) 

    # Convert predicted ratings to dataframe
    preds_df = pd.DataFrame(all_user_predicted_ratings, columns = pivot_df.columns)
    a = recommend_items(userID, pivot_df, preds_df, 8)
    return a


@app.route("/")
def index():
    return render_template("index.html")

@app.route("/login",methods=['POST','GET'])
def login():
    return render_template("login.html")


#OTP IMPLEMENT



fname = ''
lname = ''
country = ''
city_name = ''
email = ''
passworde = ''
otp_genr = 0


@app.route("/otp",methods=['POST','GET'])
def otp():
    if request.method == 'POST':
        otp_entered = int(request.form.get('otp_val'))

        #otp_genr = request.args.get('otp_value')
    global otp_genr
    if otp_genr == otp_entered:
        """global email
        global passworde
        global fname
        global lname
        global country
        global city_name"""
        workbook = load_workbook(filename="input/user_details.xlsx")
        sheet = workbook.active
        pd1 = pd.read_excel('input/user_details.xlsx')
        index = int(list(pd1['index'])[-1])
        st = str(list(pd1['user_id'])[-1])
        ind = st.index('_')
        num = int(st[ind+1:])
        user_id = 'user'+'_'+str(num+1)
        values = str(index+3)
        a = "A"+values
        b = "B"+values
        c = "C"+values
        d = "D"+values
        e = "E"+values
        f = "F"+values
        g = "G"+values

        sheet[a]= index+1
        sheet[b] = user_id
        sheet[c] = email
        sheet[d] = passworde
        sheet[e] = fname+" "+lname
        sheet[f] = country
        sheet[g] = city_name
        workbook.save(filename="input/user_details.xlsx")
        workbook = load_workbook(filename="input/new_userhotelrating_small.xlsx")
        sheet = workbook.active
        pd2 = pd.read_excel('input/new_userhotelrating_small.xlsx')
        user_id_list = list(pd2['user'])
        user_id_index = len(user_id_list)+2

        hotel_id_list = ['hotel_558','hotel_603','hotel_610','hotel_574','hotel_570','hotel_587','hotel_593','hotel_605','hotel_559']
        for hotel_id in hotel_id_list:
            sheet["A"+str(user_id_index)] = user_id
            sheet["B"+str(user_id_index)] = hotel_id
            sheet["C"+str(user_id_index)] = 0
            user_id_index= int(user_id_index)+1
        workbook.save(filename="input/new_userhotelrating_small.xlsx")


        #hotel_id_index = pd1['product_code']
       

        return render_template("login.html",message="Email ID is verified. Please login")

    else:
        return render_template("otp_verify.html",message = "OTP ENTERED IS NOT VALID! PLEASE TRY AGAIN")


@app.route("/login2",methods=['POST','GET'])
def login2():
    if request.method == 'POST':
        #global variables
        global fname
        fname = request.form.get('fname')
        global lname
        lname = request.form.get('lname')
        global country
        country = request.form.get('country')
        global city_name
        city_name = request.form.get('city')
        global email
        email = request.form.get('email')
        global passworde
        passworde = request.form.get('pass')
        confirm_pass = request.form.get('conpass')
    df=pd.read_excel('input/user_details.xlsx')
    email_lis = list(df['email_id'])
    if passworde!=confirm_pass:
        return render_template('signup.html',message= 'Password and Confirm Password must be same')
    if email in email_lis:
        return render_template('signup.html',message = 'Email id already registered')
    else:
        otpTablePath="input/otptable.xlsx"
        #global variable
        global otp_genr
        otp_genr=int(random.SystemRandom().randint(100000,999999))
        #print(otp_genr)
        
        mail_content = '''<body style="background-color:lightblue;width:78%;padding-left:10px;padding-right:15px"><br><br>
        <h2 style="color:blue;"> Verify Your Account </h2> <br>
        To verify your email address, please use the following One Time Password (OTP): <br>
        <h3 style="color:red;"> '''
        mail_content+= str(otp_genr)
        mail_content+= ''' </h3>
        Do not share this OTP with anyone. <br>
        Thank You. <br><br>
        <b>NOTE: Expires in 2 Minutes!!</b> <br><br> </body> 
        '''
        sender_address = 'donotreply.hotelrecomendation@gmail.com'
        sender_pass = 'Hotel@123'
        receiver_address = email

        message = MIMEMultipart()
        message['From'] = sender_address
        message['To'] = receiver_address

        message['Subject'] = 'OTP VERIFICATION'
        message.attach(MIMEText(mail_content, 'html'))
        session = smtplib.SMTP('smtp.gmail.com', 587)
        session.starttls()
        session.login(sender_address, sender_pass)
        text = message.as_string()
        session.sendmail(sender_address, receiver_address, text)
        session.quit()
        print('Mail Sent')
        
       


        
        
        return render_template("otp_verify.html")

@app.route("/signup",methods=['POST','GET'])
def signup():
    return render_template("signup.html")

@app.route('/logout')
def logout():
   # remove the username from the session if it is there
   session.pop('username', None)
   print('LOGGED OUT!!!')
   return render_template('index.html')

uname = ''


hotels1 = []
hotels = []
@app.route("/hotelrecommendpage")
def hotelrecommendpage():
    return render_template("hotel_recommend.html",hotels1=hotels1[:6],hotels = hotels,name=uname)
  

@app.route("/success",methods=['POST','GET'])   
def success():
    return render_template("final.html",name = uname)


@app.route("/payment")
def payment():
    if request.method == 'GET':
        cost_and_cur = request.args.get('hotelcost')
        no_of_people = request.args.get('pple')
    indi = cost_and_cur.index('.')
    cost = int(cost_and_cur[:indi])
    total_cost = str(int(cost)*(int(no_of_people)))+' '+str(cost_and_cur[indi+3:])


    return render_template("payment.html",cost=total_cost)

index_fin_val = 0   
@app.route("/exist",methods=['POST','GET'])
def exist():
    if request.method == 'POST':
        user_name = request.form.get('uname')
        password = request.form.get('pass')
        global uname

        global email
        email = user_name
        pd1 = pd.read_excel('input/user_details.xlsx')
        mail= list(map(str,pd1['email_id']))
        passw = list(map(str,pd1['password']))
        name = list(map(str,pd1['name']))
        user_index = list(map(str,pd1['user_id']))
        index_list = list(map(int,pd1['index']))
        if user_name in mail:
            i = mail.index(user_name)
            if passw[i] == password:
                global index_fin_val
                index_fin_val = index_list[i]
                uname = name[i]
                session['username'] = uname
                if 'username' in session:
                    uname = name[i]
                    print('logged in as',uname)
                    return render_template("new.html",name = name[i])
                
                return "You are not logged in <br><a href = '/templates/login.html'></b>" + \
            "click here to log in</b></a>"
                
            else:
                return render_template('login.html',message='WRONG PASSWORD. PLEASE TRY AGAIN !!')
        
        else:
            return render_template('login.html',message = 'EMAIL ID DOES NOT EXIST')

    else:
        return render_template("new.html",name = uname)

    
    
@app.route('/display' ,methods=['POST','GET'])
def display():
    if request.method == 'GET':
        hotel_name = request.args.get('h_name')
    result = []
    df=pd.read_excel('input/new_hotel_data.xlsx')
    for obj in df.values:
        if obj[1]== hotel_name:
            return render_template("display.html",hotel=obj,currency=obj[11],name=uname)
            


@app.route('/exist_result' ,methods=['POST','GET'])
def exist_result():
    data =[]
    if request.method == 'GET':
        city = request.args.get('q1')
        sort_option = request.args.get('q2')
        #index = request.args.get('indexval')
    flag = True
    ##filtering
    global email
    global hotels1
    global hotels
    global index_fin_val
    hotels1 = []
    df2 = pd.read_excel('input/new_hotel_data.xlsx')
    df=pd.read_excel('input/user_details.xlsx')
    df1 = pd.read_excel('input/new_userhotelrating_small.xlsx')
    #user_id from user_details
    email_list = list(df['email_id'])
    email_index = email_list.index(email)
    user_index_list = list(df['user_id'])
    user_id_to_s = user_index_list[email_index]

    user_id_from_rat = list(df1['user'])
    rating_list = list(df1['rating'])
    user_id_from_rat_index = user_id_from_rat.index(user_id_to_s)
    user_id_rating_list = []
    for t in range(user_id_from_rat_index,len(user_id_from_rat)):
        if user_id_from_rat[t] == user_id_to_s:
            user_id_rating_list.append(rating_list[t])

    
    for k in range(len(user_id_rating_list)):
        if user_id_rating_list[k]==0:
            pass
        else:
            flag = False
    #print('rating list',user_id_rating_list)
    if flag == True:
        hotels = []
    else:

        data = int(index_fin_val)
        recomen_pro = recom(int(data))  #function call recom
            # a = recom(1,0)
        l=recomen_pro.index[0:6]
        res = list(l)
        #print(res)
                # break
        df=pd.read_excel('input/new_hotel_data.xlsx')
        
        hotels = []
                
        for k in res:
            for i,j in df.iterrows():
                a = j[0]
                a = str(a)
                a = a.strip()
                if a == k:  
                    a = j['hotel name'],j['City']
                    hotels.append(a)

        print('Output from recommendation algorithm',hotels)
   
        

   
    print(sort_option,city)
    for index,rows in df2.iterrows():
        if rows['City']==city:
            hotels1.append((rows['hotel name'],rows['City'],rows['Average Cost for two'],rows['Aggregate rating'],rows['Currency']))
    if sort_option == "rating":
        #hotels1 = hotels1.sort(key = lambda x: x[2],reverse = True)  
        hotels1 = sorted(hotels1, key = lambda x: x[3],reverse=True)
       
    if sort_option == "costhtl":
        hotels1 = sorted(hotels1,key = lambda x: x[2],reverse = True)  
    if sort_option == "costlth":
        hotels1 = sorted(hotels1, key = lambda x: x[2])
    print('Filtered hotels',hotels1[:6])
    ##colloborative algo
    #pd2 = pd.read_excel('input/new_userhotelrating_small.xlsx')
    return render_template("hotel_recommend.html",hotels1=hotels1[:6],hotels = hotels,name = uname)
    
    
    


if __name__ == '__main__':
    app.run(debug=True,threaded=False)
